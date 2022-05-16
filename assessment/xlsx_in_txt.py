from openpyxl import load_workbook

wb = load_workbook('popisyat.xlsx')
sheet = wb['Лист1']
# print(sheet.max_row)

list_text = []
for i in range(1, sheet.max_row + 1):
    list_text.append(sheet.cell(row=i, column=1).value)
    # print(i, sheet.cell(row=i, column=1).value)
# print(list_text)

f = open("test.txt", "w")
f.write(str(list_text))
f.close()