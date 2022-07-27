from openpyxl import load_workbook
# модуль чтения содержимого xlsx и записи данных в список list_text для дальнейшей работы

def read_data():
    wb = load_workbook('popisyat.xlsx') # читаем xlsx
    sheet = wb['Лист1'] # активируем нужный нам Лист

    list_text = [] # создаем список для записи данных из xlsx
    for i in range(1, sheet.max_row + 1):  # для каждой строки из выбранного Листа
        list_text.append(sheet.cell(row=i, column=1).value)  # читаем значение первого столбца и добавляем его в список результатов
    return list_text  # на выходе получаем список с данными
